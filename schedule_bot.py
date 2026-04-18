import logging
import openpyxl
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, ReplyKeyboardMarkup, KeyboardButton
from telegram.ext import Application, CommandHandler, MessageHandler, filters, ContextTypes, ConversationHandler, CallbackQueryHandler
from datetime import datetime, timedelta, time
import os
import re
import shutil

# Настройка логирования
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# Состояния для ConversationHandler
SELECT_GROUP, SETTINGS, SEARCH_DATE, SEARCH_NAME = range(4)

# Маппинг русских месяцев
MONTHS_RU = {
    'январь': 1, 'февраль': 2, 'март': 3, 'апрель': 4, 'май': 5, 'июнь': 6,
    'июль': 7, 'август': 8, 'сентябрь': 9, 'октябрь': 10, 'ноябрь': 11, 'декабрь': 12
}

# Известные группы
KNOWN_GROUPS = ['20-21', '20-22', '20-23', '11-21', '26-21', '7-21', '8-21', '8и-21', '20и-21', '20и-22']

class ScheduleBot:
    def __init__(self):
        self.excel_file = None
        self.workbook = None
        self.schedule_data = {}
        self.groups = {}
        self.user_settings = {}
        self.excel_loaded = False
        self.load_excel_on_startup()
    
    def load_excel_on_startup(self):
        """Загружает Excel файл при старте бота"""
        try:
            work_dir = os.getcwd()
            files = [f for f in os.listdir(work_dir) if f.endswith(('.xlsx', '.xls')) and not f.startswith('temp_')]
            
            if not files:
                logger.warning("⚠️ Excel файл не найден!")
                return False
            
            self.excel_file = files[0]
            file_path = os.path.join(work_dir, self.excel_file)
            logger.info(f"Найден файл: {file_path}")
            
            self.workbook = openpyxl.load_workbook(file_path, data_only=True)
            self.parse_schedule()
            self.excel_loaded = True
            logger.info(f"✅ Загружено: {len(self.groups)} групп")
            
            # Выводим статистику для отладки
            for g in sorted(self.groups.keys()):
                dates_count = len(self.schedule_data.get(g, {}))
                total_pairs = sum(len(pairs) for pairs in self.schedule_data.get(g, {}).values())
                logger.info(f"  {g}: {dates_count} дат, {total_pairs} пар")
            
            return True
            
        except Exception as e:
            logger.error(f"❌ Ошибка загрузки: {e}")
            import traceback
            logger.error(traceback.format_exc())
            self.excel_loaded = False
            return False
    
    def save_uploaded_excel(self, file_path):
        """Сохраняет загруженный Excel файл"""
        try:
            work_dir = os.getcwd()
            old_file = os.path.join(work_dir, "schedule.xlsx")
            
            if os.path.exists(old_file):
                os.remove(old_file)
            
            new_path = os.path.join(work_dir, "schedule.xlsx")
            shutil.copy2(file_path, new_path)
            
            try:
                os.remove(file_path)
            except:
                pass
            
            self.excel_file = "schedule.xlsx"
            self.workbook = openpyxl.load_workbook(new_path, data_only=True)
            self.parse_schedule()
            self.excel_loaded = True
            
            logger.info(f"✅ Обработано: {len(self.groups)} групп")
            
            # Выводим статистику
            for g in sorted(self.groups.keys()):
                dates_count = len(self.schedule_data.get(g, {}))
                total_pairs = sum(len(pairs) for pairs in self.schedule_data.get(g, {}).values())
                logger.info(f"  {g}: {dates_count} дат, {total_pairs} пар")
            
            return True
            
        except Exception as e:
            logger.error(f"❌ Ошибка сохранения: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return False
    
    def parse_schedule(self):
        """Парсит расписание из всех листов"""
        self.schedule_data = {}
        self.groups = {}
        
        for sheet_name in self.workbook.sheetnames:
            # Пропускаем служебные листы
            if sheet_name in ['Планер']:
                continue
                
            sheet = self.workbook[sheet_name]
            logger.info(f"Обработка листа: '{sheet_name}'")
            self._parse_sheet(sheet, sheet_name)
        
        logger.info(f"ИТОГО: {len(self.groups)} групп")
    
    def _parse_sheet(self, sheet, sheet_name):
        """Парсит лист с расписанием - УПРОЩЕННЫЙ МЕТОД"""
        logger.info(f"=== ЛИСТ: '{sheet.title}' ===")
        
        # Определяем группы на этом листе
        sheet_groups = self._get_groups_from_sheet_name(sheet_name)
        logger.info(f"Группы листа: {sheet_groups}")
        
        # Инициализируем данные для групп
        for g in sheet_groups:
            if g not in self.groups:
                self.groups[g] = g
            if g not in self.schedule_data:
                self.schedule_data[g] = {}
        
        # Находим все строки с группами
        group_rows = {}
        for r in range(1, min(500, sheet.max_row + 1)):
            cell_val = sheet.cell(r, 1).value
            if cell_val:
                cell_str = str(cell_val).strip()
                for g in sheet_groups:
                    if g in cell_str:
                        if g not in group_rows:
                            group_rows[g] = []
                        group_rows[g].append(r)
        
        logger.info(f"Найдены строки групп: { {g: len(rows) for g, rows in group_rows.items()} }")
        
        # Для каждой группы собираем данные
        for group, rows in group_rows.items():
            logger.info(f"Обработка группы {group}, строк: {len(rows)}")
            self._parse_group_data(sheet, group, rows)
    
    def _parse_group_data(self, sheet, group, group_rows):
        """Парсит данные для конкретной группы"""
        current_date = None
        current_month = None
        current_year = 2026
        
        for row in group_rows:
            # Ищем дату выше этой строки
            date_info = self._find_date_above(sheet, row)
            if date_info:
                current_date, current_month = date_info
                if current_month in [1, 2, 3, 4, 5, 6]:
                    current_year = 2026
                else:
                    current_year = 2025
            
            if not current_date:
                continue
            
            date_obj = datetime(current_year, current_month, current_date)
            
            # Парсим занятия в этой строке
            pairs = self._parse_row_pairs(sheet, row, date_obj)
            
            if pairs:
                if date_obj not in self.schedule_data[group]:
                    self.schedule_data[group][date_obj] = []
                
                for pair in pairs:
                    # Проверяем на дубликаты
                    exists = False
                    for p in self.schedule_data[group][date_obj]:
                        if (p.get('pair_num') == pair['pair_num'] and 
                            p.get('day') == pair['day']):
                            exists = True
                            break
                    
                    if not exists:
                        self.schedule_data[group][date_obj].append(pair)
                        logger.info(f"  {group}: {date_obj.strftime('%d.%m.%Y')} {pair['day']}-{pair['pair_num']} {pair['subject'][:30]}")
    
    def _find_date_above(self, sheet, row):
        """Ищет дату выше указанной строки"""
        for r in range(row - 1, max(1, row - 30), -1):
            cell_val = sheet.cell(r, 1).value
            if cell_val and isinstance(cell_val, str):
                val_lower = cell_val.lower().strip()
                if val_lower in MONTHS_RU:
                    # Ищем число
                    for c in [3, 4, 5]:
                        day_val = sheet.cell(r, c).value
                        if day_val:
                            try:
                                day = int(day_val)
                                month = MONTHS_RU[val_lower]
                                return (day, month)
                            except:
                                pass
        return None
    
    def _parse_row_pairs(self, sheet, row, date_obj):
        """Парсит занятия из строки и следующих за ней"""
        pairs = []
        
        # Колонки для пар
        pair_columns = [
            (2, 'Пн', 1), (4, 'Пн', 2), (6, 'Пн', 3),
            (8, 'Вт', 1), (10, 'Вт', 2), (12, 'Вт', 3),
            (14, 'Ср', 1), (16, 'Ср', 2), (18, 'Ср', 3),
            (20, 'Чт', 1), (22, 'Чт', 2), (24, 'Чт', 3),
            (26, 'Пт', 1), (28, 'Пт', 2), (30, 'Пт', 3),
            (32, 'Сб', 1), (34, 'Сб', 2), (36, 'Сб', 3),
        ]
        
        # Ищем строки с данными (обычно следующие 1-3 строки)
        for offset in range(1, 5):
            data_row = row + offset
            if data_row > sheet.max_row:
                break
            
            # Проверяем, не началась ли следующая группа
            first_cell = sheet.cell(data_row, 1).value
            if first_cell and any(g in str(first_cell) for g in self.groups.keys()):
                break
            
            for col, day_name, pair_num in pair_columns:
                if col > sheet.max_column:
                    continue
                
                cell_val = sheet.cell(data_row, col).value
                if not cell_val:
                    continue
                
                cell_str = str(cell_val).strip()
                if cell_str in ['', 'None', '-', 'СР', 'Выходной', 'Праздник', 'Наряд']:
                    continue
                
                # Определяем тип строки по содержимому
                if offset == 1:
                    # Первая строка - обычно тип занятия или дополнительная информация
                    if re.match(r'^\d+\s*\d*\s*[лпзсгкв]', cell_str.lower()):
                        # Это тип занятия
                        pair_type = self._extract_type(cell_str)
                        # Ищем предмет в следующей строке
                        subject = self._find_subject(sheet, data_row + 1, col)
                        room = self._find_room(sheet, data_row + 2, col)
                        
                        if subject:
                            pairs.append({
                                'subject': subject,
                                'room': room,
                                'type': pair_type,
                                'pair_num': pair_num,
                                'day': day_name
                            })
                else:
                    # Может быть предметом
                    if len(cell_str) > 2 and not cell_str[0].isdigit():
                        # Похоже на название предмета
                        subject = cell_str
                        room = self._find_room(sheet, data_row + 1, col)
                        pair_type = 'л'  # по умолчанию
                        
                        pairs.append({
                            'subject': subject,
                            'room': room,
                            'type': pair_type,
                            'pair_num': pair_num,
                            'day': day_name
                        })
        
        return pairs
    
    def _extract_type(self, type_str):
        """Извлекает тип занятия из строки"""
        tl = type_str.lower()
        if 'пз' in tl:
            return 'пз'
        elif 'с' in tl and 'вси' not in tl:
            return 'с'
        elif 'гз' in tl:
            return 'гз'
        elif 'кр' in tl:
            return 'кр'
        elif 'экз' in tl:
            return 'экз'
        elif 'з/о' in tl or 'зач' in tl:
            return 'з/о'
        elif 'вси' in tl:
            return 'вси'
        elif 'гу' in tl:
            return 'гу'
        return 'л'
    
    def _find_subject(self, sheet, row, col):
        """Ищет название предмета"""
        if row <= sheet.max_row and col <= sheet.max_column:
            val = sheet.cell(row, col).value
            if val:
                val_str = str(val).strip()
                if val_str and val_str not in ['', 'None', '-']:
                    return val_str
        return ''
    
    def _find_room(self, sheet, row, col):
        """Ищет аудиторию"""
        if row <= sheet.max_row and col <= sheet.max_column:
            val = sheet.cell(row, col).value
            if val:
                val_str = str(val).strip()
                if val_str and val_str not in ['', 'None', '-']:
                    return val_str
        return ''
    
    def _get_groups_from_sheet_name(self, sheet_name):
        """Определяет группы по названию листа"""
        groups = []
        
        if '20,20и' in sheet_name:
            groups = ['20-21', '20-22', '20-23', '11-21', '20и-21', '20и-22']
        elif '26' in sheet_name:
            groups = ['26-21']
        elif '7,8,8и' in sheet_name:
            groups = ['7-21', '8-21', '8и-21']
        
        return groups
    
    def get_schedule_for_group(self, group, target_date=None):
        """Получает расписание для группы"""
        if not self.excel_loaded:
            return None
        
        if target_date is None:
            target_date = datetime.now()
        
        group = str(group).strip()
        
        # Прямой поиск
        if group in self.schedule_data:
            data = self.schedule_data[group]
        else:
            # Поиск по частичному совпадению
            found = False
            for g in self.schedule_data:
                if group.lower() in g.lower() or g.lower() in group.lower():
                    data = self.schedule_data[g]
                    group = g
                    found = True
                    break
            if not found:
                logger.warning(f"Группа {group} не найдена в данных")
                return None
        
        result = []
        for date_key, pairs in data.items():
            if isinstance(date_key, datetime) and date_key.date() == target_date.date():
                result.extend(pairs)
        
        # Сортируем
        day_order = {'Пн': 0, 'Вт': 1, 'Ср': 2, 'Чт': 3, 'Пт': 4, 'Сб': 5}
        result.sort(key=lambda x: (day_order.get(x.get('day', 'Пн'), 0), x.get('pair_num', 0)))
        
        return result
    
    def get_schedule_for_days(self, group, days=2):
        """Получает расписание на несколько дней вперед"""
        if not self.excel_loaded:
            return {}
        
        result = {}
        today = datetime.now()
        
        for i in range(days):
            target_date = today + timedelta(days=i)
            schedule = self.get_schedule_for_group(group, target_date)
            if schedule:
                result[target_date] = schedule
        
        return result
    
    def find_pair_by_name(self, group, name):
        """Ищет пару по названию"""
        if not self.excel_loaded:
            return None
        
        # Ищем группу
        target_group = None
        if group in self.schedule_data:
            target_group = group
        else:
            for g in self.schedule_data:
                if group.lower() in g.lower() or g.lower() in group.lower():
                    target_group = g
                    break
        
        if not target_group:
            logger.warning(f"Группа {group} не найдена для поиска")
            return None
        
        name_lower = name.lower().strip()
        results = []
        
        logger.info(f"Поиск '{name}' в группе {target_group}")
        logger.info(f"Всего дат: {len(self.schedule_data[target_group])}")
        
        for date_obj, pairs in self.schedule_data[target_group].items():
            for pair in pairs:
                subject = pair.get('subject', '')
                if name_lower in subject.lower():
                    results.append({'date': date_obj, 'pair': pair})
                    logger.info(f"Найдено: {date_obj.strftime('%d.%m.%Y')} - {subject}")
        
        results.sort(key=lambda x: x['date'])
        return results
    
    def get_upcoming_exams(self, group, days_ahead=60):
        """Находит ближайшие зачёты/экзамены"""
        if not self.excel_loaded:
            return []
        
        target_group = None
        if group in self.schedule_data:
            target_group = group
        else:
            for g in self.schedule_data:
                if group.lower() in g.lower() or g.lower() in group.lower():
                    target_group = g
                    break
        
        if not target_group:
            return []
        
        today = datetime.now()
        exams = []
        
        for date_obj, pairs in self.schedule_data[target_group].items():
            if not isinstance(date_obj, datetime):
                continue
            
            days_diff = (date_obj.date() - today.date()).days
            if 0 <= days_diff <= days_ahead:
                for pair in pairs:
                    pt = pair.get('type', '')
                    subj = pair.get('subject', '').lower()
                    if pt in ['экз', 'з/о', 'кр', 'зач'] or 'экзамен' in subj or 'зачёт' in subj:
                        exams.append({
                            'date': date_obj,
                            'subject': pair.get('subject', ''),
                            'type': pt,
                            'days_until': days_diff
                        })
        
        return sorted(exams, key=lambda x: (x['days_until'], x['date']))
    
    def get_all_subjects(self, group):
        """Получает список всех предметов для группы"""
        if not self.excel_loaded:
            return []
        
        target_group = None
        if group in self.schedule_data:
            target_group = group
        else:
            for g in self.schedule_data:
                if group.lower() in g.lower() or g.lower() in group.lower():
                    target_group = g
                    break
        
        if not target_group:
            return []
        
        subjects = set()
        for date_obj, pairs in self.schedule_data[target_group].items():
            for pair in pairs:
                subject = pair.get('subject', '')
                if subject and subject not in ['СР', 'Выходной', 'Праздник', 'Наряд']:
                    subjects.add(subject)
        
        return sorted(subjects)


# Создаем экземпляр бота
bot = ScheduleBot()


def get_main_keyboard():
    """Создает главную клавиатуру"""
    keyboard = [
        [KeyboardButton("📅 Сегодня"), KeyboardButton("📆 На 2 дня")],
        [KeyboardButton("🔍 По дате"), KeyboardButton("🔎 По предмету")],
        [KeyboardButton("⚙️ Группа"), KeyboardButton("📊 Экзамены")],
        [KeyboardButton("📁 Загрузить Excel"), KeyboardButton("📋 Все предметы")]
    ]
    return ReplyKeyboardMarkup(keyboard, resize_keyboard=True)


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик команды /start"""
    user_id = update.effective_user.id
    
    if str(user_id) not in bot.user_settings:
        bot.user_settings[str(user_id)] = {
            'group': '20-21',
            'days_ahead': 2,
            'enabled': True
        }
    
    settings = bot.user_settings.get(str(user_id), {})
    current_group = settings.get('group', '20-21')
    
    status = "✅ Загружено" if bot.excel_loaded else "⚠️ Ожидание файла"
    groups_count = len(bot.groups)
    
    welcome_text = f"""
🎓 *Бот расписания ВУНЦ ВВС*

👤 Группа: *{current_group}*
📊 Статус: {status}
📋 Групп загружено: *{groups_count}*

📌 *Команды:*
• 📅 Сегодня - расписание на сегодня
• 📆 На 2 дня - на сегодня и завтра
• 🔍 По дате - поиск по дате
• 🔎 По предмету - поиск по названию
• ⚙️ Группа - сменить группу
• 📊 Экзамены - ближайшие зачёты/экзамены
• 📁 Загрузить Excel - загрузить файл
• 📋 Все предметы - список предметов группы
"""
    
    await update.message.reply_text(welcome_text, parse_mode='Markdown', reply_markup=get_main_keyboard())


async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик загрузки Excel файла"""
    document = update.message.document
    
    if not document.file_name.endswith(('.xlsx', '.xls')):
        await update.message.reply_text("❌ Нужен Excel файл (.xlsx или .xls)")
        return
    
    msg = await update.message.reply_text("⏳ Скачиваю и обрабатываю...")
    
    try:
        file = await context.bot.get_file(document.file_id)
        
        temp_dir = "/tmp" if os.path.exists("/tmp") else "."
        safe_name = "".join(c for c in document.file_name if c.isalnum() or c in '._-')
        temp_path = os.path.join(temp_dir, f"temp_{safe_name}")
        
        await file.download_to_drive(temp_path)
        
        if bot.save_uploaded_excel(temp_path):
            groups_count = len(bot.groups)
            groups_list = ", ".join(sorted(bot.groups.keys()))
            
            # Получаем статистику для отладки
            stats = []
            for g in sorted(bot.groups.keys()):
                dates = len(bot.schedule_data.get(g, {}))
                pairs = sum(len(p) for p in bot.schedule_data.get(g, {}).values())
                stats.append(f"{g}: {dates} дат, {pairs} пар")
            
            stats_text = "\n".join(stats[:5])
            
            await msg.edit_text(
                f"✅ *Файл загружен!*\n\n"
                f"👥 Групп: *{groups_count}*\n"
                f"📋 `{groups_list}`\n\n"
                f"📊 *Статистика:*\n`{stats_text}`",
                parse_mode='Markdown'
            )
            await update.message.reply_text(
                "✅ Готово! Используйте кнопки для работы с расписанием.",
                reply_markup=get_main_keyboard()
            )
        else:
            await msg.edit_text("❌ Ошибка обработки файла")
            
    except Exception as e:
        logger.error(f"Ошибка загрузки: {e}")
        await msg.edit_text(f"❌ Ошибка: {str(e)[:100]}")


async def show_today(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Показывает расписание на сегодня"""
    if not bot.excel_loaded:
        await update.message.reply_text("⚠️ *Сначала загрузите Excel файл!*", parse_mode='Markdown')
        return
    
    user_id = update.effective_user.id
    settings = bot.user_settings.get(str(user_id), {})
    group = settings.get('group', '20-21')
    
    target_date = datetime.now()
    schedule = bot.get_schedule_for_group(group, target_date)
    
    days_ru = ['пн', 'вт', 'ср', 'чт', 'пт', 'сб', 'вс']
    day_name = days_ru[target_date.weekday()]
    
    header = f"📅 *{target_date.strftime('%d.%m.%Y')} ({day_name})*\n👥 *{group}*\n\n"
    
    if not schedule:
        await update.message.reply_text(header + "😴 *Нет занятий*", parse_mode='Markdown')
        return
    
    text = header
    current_day = None
    
    for item in schedule:
        day = item.get('day', '')
        if day != current_day:
            if current_day is not None:
                text += "\n"
            text += f"📌 *{day}*\n"
            current_day = day
        
        emoji = {'л': '📖', 'пз': '✏️', 'с': '🗣️', 'гз': '👥', 
                'кр': '📝', 'экз': '📋', 'з/о': '✅', 'вси': '🎯',
                'гу': '📊'}.get(item.get('type', 'л'), '📚')
        pair_num = item.get('pair_num', '?')
        subject = item.get('subject', '—')
        room = item.get('room', '')
        room_text = f" 📍{room}" if room else ""
        
        text += f"{emoji} *{pair_num}* — {subject}{room_text}\n"
    
    await update.message.reply_text(text, parse_mode='Markdown')


async def show_two_days(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Показывает расписание на 2 дня"""
    if not bot.excel_loaded:
        await update.message.reply_text("⚠️ *Сначала загрузите Excel файл!*", parse_mode='Markdown')
        return
    
    user_id = update.effective_user.id
    settings = bot.user_settings.get(str(user_id), {})
    group = settings.get('group', '20-21')
    
    schedule_dict = bot.get_schedule_for_days(group, 2)
    
    if not schedule_dict:
        await update.message.reply_text(f"😴 *Нет занятий на ближайшие 2 дня*\n👥 Группа: *{group}*", parse_mode='Markdown')
        return
    
    text = f"📆 *Расписание на 2 дня*\n👥 Группа: *{group}*\n"
    
    days_ru = ['пн', 'вт', 'ср', 'чт', 'пт', 'сб', 'вс']
    
    for date_obj, schedule in schedule_dict.items():
        date_str = date_obj.strftime('%d.%m.%Y')
        day_name = days_ru[date_obj.weekday()]
        text += f"\n📅 *{date_str} ({day_name})*\n"
        
        if not schedule:
            text += "  😴 Нет занятий\n"
            continue
        
        current_day = None
        for item in schedule:
            day = item.get('day', '')
            if day != current_day:
                text += f"  📌 {day}\n"
                current_day = day
            
            emoji = {'л': '📖', 'пз': '✏️', 'с': '🗣️', 'гз': '👥',
                    'кр': '📝', 'экз': '📋', 'з/о': '✅'}.get(item.get('type', 'л'), '📚')
            pair_num = item.get('pair_num', '?')
            subject = item.get('subject', '—')
            room = item.get('room', '')
            room_text = f" 📍{room}" if room else ""
            
            text += f"  {emoji} *{pair_num}* — {subject}{room_text}\n"
    
    if len(text) > 4000:
        parts = [text[i:i+4000] for i in range(0, len(text), 4000)]
        for part in parts:
            await update.message.reply_text(part, parse_mode='Markdown')
    else:
        await update.message.reply_text(text, parse_mode='Markdown')


async def search_by_date_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Начинает поиск по дате"""
    if not bot.excel_loaded:
        await update.message.reply_text("⚠️ Сначала загрузите Excel файл!")
        return ConversationHandler.END
    
    await update.message.reply_text(
        "📅 *Введите дату в формате:*\n`ДД.ММ.ГГГГ`\n\n"
        "Например: `12.01.2026`",
        parse_mode='Markdown'
    )
    return SEARCH_DATE


async def search_by_date_handle(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обрабатывает поиск по дате"""
    query = update.message.text.strip()
    
    try:
        target_date = datetime.strptime(query, '%d.%m.%Y')
    except ValueError:
        await update.message.reply_text(
            "❌ Неверный формат. Используйте: `ДД.ММ.ГГГГ`",
            parse_mode='Markdown'
        )
        return ConversationHandler.END
    
    user_id = update.effective_user.id
    settings = bot.user_settings.get(str(user_id), {})
    group = settings.get('group', '20-21')
    
    schedule = bot.get_schedule_for_group(group, target_date)
    
    days_ru = ['пн', 'вт', 'ср', 'чт', 'пт', 'сб', 'вс']
    day_name = days_ru[target_date.weekday()]
    
    if not schedule:
        await update.message.reply_text(
            f"❌ Нет занятий на {target_date.strftime('%d.%m.%Y')} ({day_name})\n👥 Группа: *{group}*",
            parse_mode='Markdown'
        )
    else:
        text = f"📅 *{target_date.strftime('%d.%m.%Y')} ({day_name})*\n👥 *{group}*\n\n"
        for item in schedule:
            emoji = {'л': '📖', 'пз': '✏️', 'с': '🗣️', 'гз': '👥',
                    'кр': '📝', 'экз': '📋', 'з/о': '✅'}.get(item.get('type', 'л'), '📚')
            pair_num = item.get('pair_num', '?')
            subject = item.get('subject', '—')
            room = item.get('room', '')
            room_text = f" 📍{room}" if room else ""
            text += f"{emoji} *{pair_num}* — {subject}{room_text}\n"
        await update.message.reply_text(text, parse_mode='Markdown')
    
    return ConversationHandler.END


async def search_by_name_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Начинает поиск по названию"""
    if not bot.excel_loaded:
        await update.message.reply_text("⚠️ Сначала загрузите Excel файл!")
        return ConversationHandler.END
    
    user_id = update.effective_user.id
    settings = bot.user_settings.get(str(user_id), {})
    group = settings.get('group', '20-21')
    
    await update.message.reply_text(
        f"🔎 *Введите название предмета:*\n\n"
        f"👥 Группа: *{group}*\n\n"
        f"Примеры: `ИАО`, `ФП`, `ТВВС`, `ИНО`",
        parse_mode='Markdown'
    )
    return SEARCH_NAME


async def search_by_name_handle(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обрабатывает поиск по названию"""
    query = update.message.text.strip()
    user_id = update.effective_user.id
    settings = bot.user_settings.get(str(user_id), {})
    group = settings.get('group', '20-21')
    
    results = bot.find_pair_by_name(group, query)
    
    if not results:
        await update.message.reply_text(
            f"❌ *{query}* не найдено\n👥 Группа: *{group}*",
            parse_mode='Markdown'
        )
    else:
        text = f"🔎 *Результаты поиска: {query}*\n👥 Группа: *{group}*\n\n"
        
        for result in results[:15]:
            date_str = result['date'].strftime('%d.%m.%Y')
            pair = result['pair']
            emoji = {'л': '📖', 'пз': '✏️', 'с': '🗣️', 'гз': '👥',
                    'кр': '📝', 'экз': '📋', 'з/о': '✅'}.get(pair.get('type', 'л'), '📚')
            text += f"📅 {date_str} ({pair.get('day', '')})\n"
            text += f"   {emoji} *{pair.get('pair_num', '?')}* — {pair.get('subject', '')}\n\n"
        
        if len(results) > 15:
            text += f"и ещё {len(results) - 15}..."
        
        await update.message.reply_text(text, parse_mode='Markdown')
    
    return ConversationHandler.END


async def cancel_search(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Отменяет поиск"""
    await update.message.reply_text("❌ Поиск отменен", reply_markup=get_main_keyboard())
    return ConversationHandler.END


async def show_exams(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Показывает ближайшие экзамены и зачёты"""
    if not bot.excel_loaded:
        await update.message.reply_text("⚠️ Сначала загрузите Excel файл!")
        return
    
    user_id = update.effective_user.id
    settings = bot.user_settings.get(str(user_id), {})
    group = settings.get('group', '20-21')
    
    exams = bot.get_upcoming_exams(group, days_ahead=60)
    
    if not exams:
        await update.message.reply_text(
            f"✅ *Нет ближайших экзаменов/зачётов*\n👥 Группа: *{group}*",
            parse_mode='Markdown'
        )
        return
    
    text = f"📊 *Ближайшие экзамены/зачёты*\n👥 Группа: *{group}*\n\n"
    
    for exam in exams[:15]:
        days_text = f"через {exam['days_until']} дн." if exam['days_until'] > 0 else "🔥 *СЕГОДНЯ!*"
        date_str = exam['date'].strftime('%d.%m.%Y')
        
        type_emoji = {'экз': '📋', 'з/о': '✅', 'кр': '📝', 'зач': '✅'}.get(exam['type'], '📌')
        text += f"{type_emoji} {date_str} ({days_text})\n"
        text += f"   *{exam['subject']}*\n\n"
    
    await update.message.reply_text(text, parse_mode='Markdown')


async def show_all_subjects(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Показывает все предметы для текущей группы"""
    if not bot.excel_loaded:
        await update.message.reply_text("⚠️ Сначала загрузите Excel файл!")
        return
    
    user_id = update.effective_user.id
    settings = bot.user_settings.get(str(user_id), {})
    group = settings.get('group', '20-21')
    
    subjects = bot.get_all_subjects(group)
    
    if not subjects:
        await update.message.reply_text(
            f"❌ Предметы не найдены\n👥 Группа: *{group}*\n\n"
            f"Проверьте логи бота - возможно, данные не загрузились.",
            parse_mode='Markdown'
        )
        return
    
    text = f"📋 *Все предметы*\n👥 Группа: *{group}*\n\n"
    
    for i, subject in enumerate(subjects, 1):
        text += f"{i}. `{subject}`\n"
    
    if len(text) > 4000:
        parts = [text[i:i+4000] for i in range(0, len(text), 4000)]
        for part in parts:
            await update.message.reply_text(part, parse_mode='Markdown')
    else:
        await update.message.reply_text(text, parse_mode='Markdown')


async def settings_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Показывает меню настроек (выбор группы)"""
    user_id = update.effective_user.id
    settings = bot.user_settings.get(str(user_id), {})
    current_group = settings.get('group', '20-21')
    
    available_groups = sorted(bot.groups.keys()) if bot.groups else KNOWN_GROUPS
    
    keyboard = []
    for i in range(0, len(available_groups), 3):
        row = []
        for g in available_groups[i:i+3]:
            label = f"✅ {g}" if g == current_group else f"👥 {g}"
            row.append(InlineKeyboardButton(label, callback_data=f"group_{g}"))
        keyboard.append(row)
    
    keyboard.append([InlineKeyboardButton("🔙 Закрыть", callback_data="close_settings")])
    
    text = f"⚙️ *Выбор группы*\n\nТекущая группа: *{current_group}*\n\nВыберите группу:"
    
    if update.callback_query:
        await update.callback_query.edit_message_text(
            text,
            parse_mode='Markdown',
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
    else:
        await update.message.reply_text(
            text,
            parse_mode='Markdown',
            reply_markup=InlineKeyboardMarkup(keyboard)
        )


async def button_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик inline кнопок"""
    query = update.callback_query
    await query.answer()
    data = query.data
    user_id = update.effective_user.id
    
    if data == "close_settings":
        await query.edit_message_text("✅ Настройки закрыты")
        return
    
    if data.startswith("group_"):
        group_code = data[6:]
        if str(user_id) not in bot.user_settings:
            bot.user_settings[str(user_id)] = {}
        bot.user_settings[str(user_id)]['group'] = group_code
        
        available_groups = sorted(bot.groups.keys()) if bot.groups else KNOWN_GROUPS
        
        keyboard = []
        for i in range(0, len(available_groups), 3):
            row = []
            for g in available_groups[i:i+3]:
                label = f"✅ {g}" if g == group_code else f"👥 {g}"
                row.append(InlineKeyboardButton(label, callback_data=f"group_{g}"))
            keyboard.append(row)
        
        keyboard.append([InlineKeyboardButton("🔙 Закрыть", callback_data="close_settings")])
        
        await query.edit_message_text(
            f"⚙️ *Выбор группы*\n\n✅ Группа изменена на *{group_code}*\n\nВыберите группу:",
            parse_mode='Markdown',
            reply_markup=InlineKeyboardMarkup(keyboard)
        )


async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик текстовых сообщений (кнопки)"""
    text = update.message.text
    
    if text == "📅 Сегодня":
        await show_today(update, context)
    elif text == "📆 На 2 дня":
        await show_two_days(update, context)
    elif text == "🔍 По дате":
        await search_by_date_start(update, context)
    elif text == "🔎 По предмету":
        await search_by_name_start(update, context)
    elif text == "⚙️ Группа":
        await settings_menu(update, context)
    elif text == "📊 Экзамены":
        await show_exams(update, context)
    elif text == "📁 Загрузить Excel":
        await update.message.reply_text("📁 Отправьте Excel файл с расписанием")
    elif text == "📋 Все предметы":
        await show_all_subjects(update, context)
    else:
        await update.message.reply_text("Используйте кнопки меню для навигации")


def main():
    """Запуск бота"""
    token = os.getenv("TELEGRAM_BOT_TOKEN")
    if not token:
        logger.error("❌ Не указан TELEGRAM_BOT_TOKEN!")
        return
    
    logger.info(f"🚀 Запуск бота. Рабочая директория: {os.getcwd()}")
    
    application = Application.builder().token(token).build()
    
    application.add_handler(CommandHandler("start", start))
    application.add_handler(CommandHandler("settings", settings_menu))
    
    application.add_handler(MessageHandler(
        filters.Document.FileExtension("xlsx") | filters.Document.FileExtension("xls"), 
        handle_document
    ))
    
    date_conv = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex(r'^🔍 По дате$'), search_by_date_start)],
        states={
            SEARCH_DATE: [MessageHandler(filters.TEXT & ~filters.COMMAND, search_by_date_handle)]
        },
        fallbacks=[CommandHandler("cancel", cancel_search)],
        allow_reentry=True
    )
    
    name_conv = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex(r'^🔎 По предмету$'), search_by_name_start)],
        states={
            SEARCH_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, search_by_name_handle)]
        },
        fallbacks=[CommandHandler("cancel", cancel_search)],
        allow_reentry=True
    )
    
    application.add_handler(date_conv)
    application.add_handler(name_conv)
    
    application.add_handler(CallbackQueryHandler(button_handler))
    
    application.add_handler(MessageHandler(
        filters.TEXT & ~filters.COMMAND, 
        handle_text
    ))
    
    logger.info("✅ Бот запущен и готов к работе!")
    application.run_polling(allowed_updates=Update.ALL_TYPES)


if __name__ == "__main__":
    main()
